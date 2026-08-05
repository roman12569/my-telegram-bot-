# ==============================================================================
# OEB NEXUS - PART 1: Core Architecture, Imports, DB, Caching & Server Engine
# ==============================================================================
# CRITICAL: This part contains ALL global imports, configurations, database 
# connections, thread executors, and caching systems. DO NOT duplicate these 
# in subsequent parts.

import os
import gc
import io
import re
import csv
import time
import random
import string
import hashlib
import threading
import collections
import concurrent.futures
from datetime import datetime, timedelta, timezone

import requests
import openpyxl
import pyotp
import telebot
from telebot import types
from telebot.apihelper import ApiTelegramException

from flask import Flask
from waitress import serve

from pymongo import MongoClient
from pymongo.errors import DuplicateKeyError, PyMongoError

# ==========================================
# 1. Environment Variables & Constants
# ==========================================
BOT_TOKEN = os.getenv("BOT_TOKEN","8765437674:AAGCMs5y3_8WXduxd_kSpF_4Jm-2EovgHl4")
MONGO_URL = os.getenv("MONGO_URL","mongodb+srv://admin:W3tcfbw_EW8QfR-@cluster0.nvv6umd.mongodb.net/?appName=Cluster0")
ADMIN_ID = int(os.getenv("ADMIN_ID", "6257034751"))
BACKUP_CHANNEL_ID = int(os.getenv("BACKUP_CHANNEL_ID", "-1003943094107"))

BD_TIMEZONE = timezone(timedelta(hours=6))

UAS = [
    "Mozilla/5.0 (Linux; Android 13; SM-G973F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 16_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.5 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 13; Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36"
]

REQUIRED_CHANNELS = [
    {"name": "OEB Official", "username": "@oeb_official", "url": "https://t.me/oeb_official"},
    {"name": "OEB Tasks", "username": "@oeb_tasks", "url": "https://t.me/oeb_tasks"}
]

# ==========================================
# 2. Bot & Optimized MongoDB Connection
# ==========================================
bot = telebot.TeleBot(BOT_TOKEN)

mongo_client = MongoClient(
    MONGO_URL,
    maxPoolSize=20,
    minPoolSize=5,
    socketTimeoutMS=10000,
    connect=False
)

db = mongo_client["earning_bazar_advanced"]
users_col = db["users"]
submissions_col = db["submissions"]
settings_col = db["settings"]
withdrawals_col = db["withdrawals"]
blacklisted_payloads_col = db["blacklisted_payloads"]
ai_logs_col = db["ai_logs"]

# Background indexes for performance
submissions_col.create_index([("track_id", 1)], unique=True, background=True)
submissions_col.create_index([("uid", 1)], background=True)
submissions_col.create_index([("date_key", 1)], background=True)
submissions_col.create_index([("status", 1)], background=True)

# ==========================================
# 3. Strict RAM Protection (Bounded Executors)
# ==========================================
class GuaranteedBoundedExecutor:
    def __init__(self, max_workers, max_queue_size=None):
        self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=max_workers)
        self.max_queue_size = max_queue_size
        self.queue_semaphore = threading.Semaphore(max_queue_size) if max_queue_size else None

    def submit(self, fn, *args, **kwargs):
        if self.queue_semaphore:
            if not self.queue_semaphore.acquire(blocking=False):
                return None
        return self.executor.submit(self._task_wrapper, fn, *args, **kwargs)

    def _task_wrapper(self, fn, *args, **kwargs):
        try:
            return fn(*args, **kwargs)
        finally:
            if self.queue_semaphore:
                self.queue_semaphore.release()

    def shutdown(self, wait=True):
        self.executor.shutdown(wait=wait)

# Initialize specific executors to prevent RAM overflow
background_executor = GuaranteedBoundedExecutor(max_workers=3, max_queue_size=100)
heavy_task_executor = GuaranteedBoundedExecutor(max_workers=2, max_queue_size=50)
live_check_executor = GuaranteedBoundedExecutor(max_workers=2, max_queue_size=100)
cache_executor = concurrent.futures.ThreadPoolExecutor(max_workers=2, thread_name_prefix="CacheSync")

# ==========================================
# 4. In-Memory Caching System (MongoDB Saver)
# ==========================================
class FastSettingsCache:
    def __init__(self):
        self.lock = threading.Lock()
        self.cache = {}
        self._load_initial_cache()

    def _load_initial_cache(self):
        try:
            for doc in settings_col.find():
                if 'key' in doc and 'value' in doc:
                    self.cache[doc['key']] = doc['value']
        except Exception as e:
            print(f"[FastSettingsCache] Init error: {e}")

    def get(self, key, default=None):
        with self.lock:
            if key in self.cache:
                return self.cache[key]
        db_doc = settings_col.find_one({"key": key})
        if db_doc and 'value' in db_doc:
            val = db_doc['value']
            with self.lock:
                if key not in self.cache:
                    self.cache[key] = val
            return val
        return default

    def set(self, key, value):
        with self.lock:
            self.cache[key] = value
        cache_executor.submit(self._async_db_update, key, value)

    def _async_db_update(self, key, value):
        try:
            settings_col.update_one({"key": key}, {"$set": {"value": value}}, upsert=True)
        except Exception as e:
            print(f"[FastSettingsCache] Async update failed: {e}")

class MongoDict:
    def __init__(self, collection, max_cache_size=2000):
        self.collection = collection
        self.max_cache_size = max_cache_size
        self.lock = threading.Lock()
        self.cache = collections.OrderedDict()

    def get(self, key, default=None):
        with self.lock:
            if key in self.cache:
                self.cache.move_to_end(key)
                return self.cache[key]
        db_doc = self.collection.find_one({"_id": key})
        if db_doc and 'value' in db_doc:
            val = db_doc['value']
            with self.lock:
                if key in self.cache:
                    self.cache.move_to_end(key)
                    return self.cache[key]
                self.cache[key] = val
                self.cache.move_to_end(key)
                if len(self.cache) > self.max_cache_size:
                    self.cache.popitem(last=False)
            return val
        return default

    def __setitem__(self, key, value):
        with self.lock:
            if key in self.cache:
                self.cache.move_to_end(key)
            self.cache[key] = value
            if len(self.cache) > self.max_cache_size:
                self.cache.popitem(last=False)
        cache_executor.submit(self._async_db_upsert, key, value)

    def pop(self, key, default=None):
        with self.lock:
            val = self.cache.pop(key, default)
        if val is not default:
            cache_executor.submit(self._async_db_delete, key)
            return val
        return default

    def _async_db_upsert(self, key, value):
        try:
            self.collection.update_one({"_id": key}, {"$set": {"value": value}}, upsert=True)
        except Exception as e:
            print(f"[MongoDict] Async upsert failed: {e}")

    def _async_db_delete(self, key):
        try:
            self.collection.delete_one({"_id": key})
        except Exception as e:
            print(f"[MongoDict] Async delete failed: {e}")

# Initialize Cache Instances
fast_settings = FastSettingsCache()
user_states = MongoDict(db['user_states'], max_cache_size=2000)

def get_setting(key, default=None):
    return fast_settings.get(key, default)

def update_setting(key, value):
    fast_settings.set(key, value)

# ==========================================
# 5. Time Utilities
# ==========================================
def get_bd_time():
    return datetime.now(BD_TIMEZONE)

def parse_iso_datetime(dt_val):
    if dt_val is None: return None
    try:
        if isinstance(dt_val, str):
            dt = datetime.fromisoformat(dt_val)
        elif isinstance(dt_val, datetime):
            dt = dt_val
        else:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=BD_TIMEZONE)
        else:
            dt = dt.astimezone(BD_TIMEZONE)
        return dt
    except (ValueError, TypeError):
        return None

# ==========================================
# 6. 24/7 Production Server (Flask + Waitress)
# ==========================================
app = Flask(__name__)

@app.route('/')
def home():
    return "OEB NEXUS Engine Active"

def run_web():
    serve(app, host='0.0.0.0', port=10000)

# NOTE: The actual execution block (starting the web thread and bot polling) 
# is intentionally omitted here and placed at the very end of Part 10 to ensure 
# all handlers are registered before the bot starts polling.



# ==============================================================================
# OEB NEXUS - PART 2: Core Utilities, Security, User Management & Nightly Daemon
# ==============================================================================
# DEPENDENCY: Assumes Part 1 globals (bot, db, collections, executors, time utils) exist.

# ==========================================
# 1. String Sanitization & ID Generators
# ==========================================
def sanitize_html(text):
    """Strictly sanitizes text for Telegram's HTML parse mode."""
    if not text:
        return "User"
    text = str(text)
    # Order is critical to prevent double-escaping
    text = text.replace("&", "&amp;")
    text = text.replace("<", "&lt;")
    text = text.replace(">", "&gt;")
    return text

def generate_tracking_id():
    """Generates a unique tracking ID for submissions."""
    timestamp = int(get_bd_time().timestamp())
    random_suffix = random.randint(100, 999)
    return f"SUB-{timestamp}-{random_suffix}"

def generate_withdraw_id():
    """Generates a unique withdrawal ID."""
    timestamp = int(get_bd_time().timestamp())
    random_suffix = random.randint(100, 999)
    return f"WDR-{timestamp}-{random_suffix}"

# ==========================================
# 2. User Data & State Management
# ==========================================
def get_user_data(chat_id):
    """Fetches user data. Creates a new document if the user doesn't exist."""
    user = users_col.find_one({"_id": chat_id})
    if not user:
        new_user = {
            "_id": chat_id,
            "username": "",
            "first_name": "Worker",
            "balance": 0.0,
            "hold_balance": 0.0,
            "banned": False,
            "custom_password": "",
            "role": "member",
            "virtual_wallet": 0.0,
            "assigned_sub_admin": None,
            "last_bonus_date": None
        }
        try:
            users_col.insert_one(new_user)
            return new_user
        except DuplicateKeyError:
            # Race condition handled: fetch the newly created document
            return users_col.find_one({"_id": chat_id})
    return user

def update_user_field(chat_id, field, val):
    """Asynchronously updates a specific field in the user's document."""
    if field == "_id":
        return
    def _db_update():
        try:
            users_col.update_one({"_id": chat_id}, {"$set": {field: val}}, upsert=True)
        except Exception as e:
            print(f"[AsyncUpdate] Failed to update {field} for {chat_id}: {e}")
    background_executor.submit(_db_update)

def is_user_banned(chat_id):
    """Checks if a user is banned."""
    user = users_col.find_one({"_id": chat_id}, {"banned": 1})
    return user.get("banned", False) if user else False

# ==========================================
# 3. Force Join Validator
# ==========================================
def check_force_join(user_id):
    """Verifies if the user has joined all required channels."""
    if user_id == ADMIN_ID:
        return True
    for channel in REQUIRED_CHANNELS:
        try:
            member = bot.get_chat_member(channel["username"], user_id)
            if member.status in ['left', 'kicked']:
                return False
        except ApiTelegramException:
            return False
        except Exception:
            return False
    return True

# ==========================================
# 4. Security, UID & Payload Validation
# ==========================================
def validate_strict_password(password, rule):
    """Validates a password against a strict suffix rule."""
    if not rule or str(rule).strip().lower() == "none":
        return True
    if not password:
        return False
    return str(password).strip().endswith(str(rule).strip())

# Pre-compile regex for maximum performance
UID_PATTERN = re.compile(
    r'(?:c_user=|id=|profile\.php\?id=|/u/)(\d{8,20})|(?<!\d)(\d{8,20})(?!\d)'
)

def extract_numeric_uid(text):
    """Extracts the first valid numeric UID (8-20 digits) from text."""
    if not text: return None
    match = UID_PATTERN.search(str(text))
    if match:
        return match.group(1) or match.group(2)
    return None

def is_duplicate_uid(uid):
    """Checks if a UID already exists in the submissions collection."""
    if not uid: return True
    try:
        return submissions_col.find_one({"uid": str(uid)}, {"_id": 1}) is not None
    except PyMongoError:
        return True  # Fail safe

def generate_payload_hash(payload):
    """Generates a strict SHA-256 hash of the payload (strips all whitespace)."""
    if not payload: return None
    clean_payload = re.sub(r'\s+', '', str(payload))
    return hashlib.sha256(clean_payload.encode('utf-8')).hexdigest()

def is_payload_blacklisted(p_hash):
    """Checks if a payload hash exists in the blacklist collection."""
    if not p_hash: return True
    try:
        return blacklisted_payloads_col.find_one({"_id": p_hash}) is not None
    except PyMongoError:
        return True

def is_valid_cookies(cookie_str):
    """Validates if the cookie string contains essential Facebook session tokens."""
    if not cookie_str: return False
    lower_cookies = str(cookie_str).lower()
    return any(token in lower_cookies for token in ["c_user=", "datr=", "xs=", "sessionid="])

# ==========================================
# 5. AI Logger & External APIs
# ==========================================
def log_ai_report(issue_type, description, fix_action):
    """Logs AI healing actions to the database and alerts the Admin asynchronously."""
    def task():
        try:
            log_timestamp = get_bd_time()
            ai_logs_col.insert_one({
                "timestamp": log_timestamp,
                "type": str(issue_type),
                "description": str(description),
                "action": str(fix_action)
            })
            alert_msg = (
                f"🧠 <b>AI AUTO-HEALING REPORT</b>\n\n"
                f"🔴 <b>Issue:</b> {sanitize_html(issue_type)}\n"
                f"📝 <b>Desc:</b> {sanitize_html(description)}\n"
                f"🛠 <b>Fix:</b> {sanitize_html(fix_action)}\n"
                f"🕒 <b>Time:</b> {log_timestamp.strftime('%Y-%m-%d %H:%M:%S')}"
            )
            if ADMIN_ID:
                bot.send_message(ADMIN_ID, alert_msg, parse_mode="HTML")
        except Exception as e:
            print(f"[AI Logger] Error: {e}")
    background_executor.submit(task)

def fetch_temp_email_data(url):
    """Production-ready helper for external Temp Email API GET requests."""
    try:
        response = requests.get(url, timeout=5.0)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        print(f"[TempEmailAPI] Error: {e}")
        return None

# ==========================================
# 6. Escrow & Nightly Cleanup Daemon
# ==========================================
last_report_date = None
last_wipe_date = None

def escrow_and_cleanup_daemon():
    """Background daemon for daily reporting and nightly database cleanup."""
    global last_report_date, last_wipe_date
    
    while True:
        try:
            time.sleep(30)
            now = get_bd_time()
            today_str = now.strftime("%Y-%m-%d")
            
            # Condition 1: Daily Report at 23:55
            if now.hour == 23 and now.minute == 55 and last_report_date != now.date():
                last_report_date = now.date()
                pipeline = [
                    {"$match": {"date_key": today_str}},
                    {"$group": {
                        "_id": "$category",
                        "count": {"$sum": 1},
                        "total_rate": {"$sum": {"$ifNull": ["$rate", 0]}}
                    }}
                ]
                results = list(submissions_col.aggregate(pipeline))
                diary_msg = f"📊 <b>DAILY DIARY - {today_str}</b>\n\n"
                total_tasks, total_earnings = 0, 0.0
                
                if not results:
                    diary_msg += "No tasks submitted today."
                else:
                    for res in results:
                        cat = res['_id'] or 'Uncategorized'
                        cnt, rate = res['count'], res['total_rate']
                        total_tasks += cnt
                        total_earnings += rate
                        diary_msg += f"▪️ <b>{cat}</b>: {cnt} tasks | 💰 {rate:.2f} BDT\n"
                    diary_msg += f"\n📈 <b>Total Tasks</b>: {total_tasks}\n💵 <b>Total Payout</b>: {total_earnings:.2f} BDT"
                
                if ADMIN_ID:
                    bot.send_message(ADMIN_ID, diary_msg, parse_mode="HTML")

            # Condition 2: Nightly Wipe at 00:00
            if now.hour == 0 and now.minute == 0 and last_wipe_date != now.date():
                last_wipe_date = now.date()
                yesterday_str = (now - timedelta(days=1)).strftime("%Y-%m-%d")
                
                # Delete ONLY Approved/Rejected tasks. Never touch "Hold" tasks.
                delete_query = {
                    "date_key": yesterday_str,
                    "status": {"$in": ["Approved", "Rejected"]}
                }
                deleted_count = submissions_col.delete_many(delete_query).deleted_count
                
                cleanup_msg = (
                    f"🧹 <b>NIGHTLY CLEANUP REPORT</b>\n\n"
                    f"🗓 <b>Date Cleared</b>: {yesterday_str}\n"
                    f"🗑 <b>Tasks Deleted</b>: {deleted_count}\n"
                    f"✅ <b>Status</b>: 'Hold' tasks preserved."
                )
                if ADMIN_ID:
                    bot.send_message(ADMIN_ID, cleanup_msg, parse_mode="HTML")
                    
        except Exception as e:
            print(f"[Daemon Error] {e}")
            time.sleep(60)

# Start the daemon in a background thread
threading.Thread(target=escrow_and_cleanup_daemon, daemon=True, name="NightlyCleanupDaemon").start()




# ==============================================================================
# OEB NEXUS - PART 3: Live Checker Engine, Dynamic Pricing & Shift Control
# ==============================================================================
# DEPENDENCY: Assumes Part 1 & 2 globals (bot, db, collections, executors, time utils, 
# security functions, UAS) exist.

# ==========================================
# 1. Facebook Live Checker
# ==========================================
def check_live_account(uid):
    """
    Checks if a Facebook UID is currently active, suspended, or cloud-blocked.
    WARNING: This function contains time.sleep(). It MUST ONLY be executed 
    inside the live_check_executor to prevent blocking the main bot thread.
    """
    clean_uid = extract_numeric_uid(uid)
    if not clean_uid:
        return False, "Invalid UID"

    # Anti-spam delay to prevent triggering Cloudflare rate limits
    time.sleep(random.uniform(0.5, 1.5))

    try:
        headers = {
            "User-Agent": random.choice(UAS),
            "Accept-Language": "en-US,en;q=0.9",
            "Sec-Fetch-Mode": "navigate"
        }
        
        target_url = f"https://m.facebook.com/profile.php?id={clean_uid}"
        
        # Strict 5.0s timeout, allow redirects to catch login/checkpoint redirects
        res = requests.get(target_url, headers=headers, timeout=5.0, allow_redirects=True)

        content = res.text.lower()
        res_url = (res.url or "").lower()

        # --- Cloud IP Block Fix ---
        if res.status_code == 403 or "login" in res_url or "checkpoint" in res_url:
            if "c_user" not in res_url and clean_uid not in res_url:
                return False, "Cloud Blocked"

        # --- Status Code Check ---
        if res.status_code != 200:
            return False, "Suspended/Dead"

        # --- Content Heuristics (Checkpoint/Dead) ---
        if 'content="no-cache"' in content or "not found" in content:
            return False, "Checkpoint"

        # --- Content Heuristics (Live) ---
        if "profile_ring" in content or "mbasic_inline_feed_composer" in content or clean_uid in res_url:
            return True, "Live"

        return False, "Suspended"

    except Exception:
        return False, "Network Error"

# ==========================================
# 2. Instagram Live Checker
# ==========================================
def check_ig_username_live(username):
    """
    Checks if an Instagram username is active. 
    Includes a graceful fallback for IG's aggressive anti-scraping mechanisms.
    WARNING: Contains time.sleep(). MUST be executed inside live_check_executor.
    """
    if not username:
        return False, "Invalid Username"

    clean_user = str(username).replace("@", "").strip()
    if not clean_user:
        return False, "Invalid Username"

    time.sleep(random.uniform(0.5, 1.5))

    try:
        headers = {"User-Agent": random.choice(UAS)}
        target_url = f"https://www.instagram.com/{clean_user}/"
        res = requests.get(target_url, headers=headers, timeout=5.0)

        res_url = (res.url or "").lower()

        if res.status_code == 403 or "login" in res_url:
            return False, "Cloud Blocked"

        if res.status_code == 200 and "page not found" not in res.text.lower():
            return True, "Live"

        return False, "Dead"

    except Exception:
        # CRITICAL BUSINESS LOGIC: Instagram aggressively blocks scrapers.
        # Assume live to prevent discarding valid accounts due to IP limits.
        return True, "Assumed Live"

# ==========================================
# 3. Dynamic Surge Pricing Engine
# ==========================================
def get_active_surge_bonus():
    """
    Calculates the current active surge bonus based on configuration.
    Safely handles timezone-aware datetime comparisons.
    """
    surge_config = get_setting("surge_pricing", {"active": False, "bonus": 0.0, "expires_at": None})
    
    if surge_config.get("active"):
        expires_at_str = surge_config.get("expires_at")
        
        if expires_at_str:
            expires_at = parse_iso_datetime(expires_at_str)
            if expires_at and get_bd_time() < expires_at:
                return float(surge_config.get("bonus", 0.0))
            return 0.0
            
        # If active but no expiry is set, the bonus applies indefinitely
        return float(surge_config.get("bonus", 0.0))
        
    return 0.0

def get_current_task_rate(cat_key):
    """
    Calculates the final payout rate for a specific task category.
    Combines the base rate from settings with any active surge bonus.
    """
    default_rates = {
        "fb_cookie": 5.0, 
        "fb_2fa": 6.0, 
        "ig_cookie": 8.0, 
        "ig_2fa": 10.0
    }
    
    rates = get_setting("rates", default_rates)
    base_rate = float(rates.get(cat_key, 5.0))
    surge_bonus = get_active_surge_bonus()
    
    return base_rate + surge_bonus

# ==========================================
# 4. Shift Control & Submission Gatekeeper
# ==========================================
def get_shift_config():
    """
    Retrieves the shift configuration, dynamically injecting today's BD date 
    into the default configuration to ensure daily shift resets.
    """
    default_config = {
        "current_date": get_bd_time().strftime("%Y-%m-%d"),
        "deadlines": {
            "fb_cookie": "21:20",
            "fb_2fa": "21:20",
            "ig_cookie": "20:20",
            "ig_2fa": "20:20",
            "default": "23:59"
        }
    }
    
    return get_setting("shift_config", default_config)

def is_submission_allowed(cat_key, req_time):
    """
    Strictly validates if a task submission is allowed based on the current 
    shift date and category-specific daily deadlines.
    """
    try:
        shift = get_shift_config()
        
        # Rule 1 (Date Check)
        req_date_str = req_time.strftime("%Y-%m-%d")
        if req_date_str != shift.get("current_date"):
            return False, "⚠️ আজকের শিফট এখনো চালু হয়নি!"
            
        # Rule 2 (Time Check)
        deadlines = shift.get("deadlines", {})
        deadline_str = deadlines.get(cat_key, deadlines.get("default", "23:59"))
        
        hours, minutes = map(int, deadline_str.split(":"))
        target_deadline = req_time.replace(hour=hours, minute=minutes, second=0, microsecond=0)
        
        if req_time > target_deadline:
            return False, f"⚠️ ডেডলাইন {deadline_str} শেষ!"
            
        return True, "Allowed"
        
    except Exception:
        # CRITICAL FALLBACK: Default to ALLOWED to prevent blocking legitimate 
        # worker submissions due to a backend configuration glitch.
        return True, "Allowed"


# ==============================================================================
# OEB NEXUS - PART 4: Smart Submission Hub (Single & Bulk Text Processing)
# ==============================================================================
# DEPENDENCY: Assumes Parts 1-3 globals exist. Handles state transitions for 
# single/bulk submissions, concurrent live checking, and atomic DB updates.

# ==========================================
# 1. Helper: Save Submission to DB & Update Balance
# ==========================================
def _save_submission(chat_id, uid, payload, category):
    """
    Atomically saves a task to MongoDB and increments the user's hold balance.
    Returns the generated track_id and rate if successful, None/0.0 if duplicate.
    """
    track_id = generate_tracking_id()
    date_key = get_bd_time().strftime("%Y-%m-%d")
    rate = get_current_task_rate(category)
    
    doc = {
        "_id": track_id,
        "track_id": track_id,
        "uid": str(uid),
        "payload": payload,
        "category": category,
        "rate": rate,
        "status": "Hold",
        "date_key": date_key,
        "submitted_at": get_bd_time(),
        "user_id": chat_id
    }
    
    try:
        submissions_col.insert_one(doc)
        # Atomic increment of hold_balance
        users_col.update_one({"_id": chat_id}, {"$inc": {"hold_balance": rate}})
        return track_id, rate
    except DuplicateKeyError:
        return None, 0.0
    except Exception as e:
        print(f"[SaveSubmission] DB Error: {e}")
        return None, 0.0

# ==========================================
# 2. Single Submission Flow Handlers
# ==========================================
def handle_single_uid(chat_id, text):
    """State: AWAITING_UID -> AWAITING_SINGLE_DATA"""
    uid = extract_numeric_uid(text)
    if not uid or is_duplicate_uid(uid):
        return "❌ Invalid or Duplicate UID. Please try again."
        
    state_data = user_states.get(chat_id, {})
    category = state_data.get("category", "fb_cookie")
    
    allowed, msg = is_submission_allowed(category, get_bd_time())
    if not allowed:
        return msg
        
    # Transition state
    state_data["state"] = "AWAITING_SINGLE_DATA"
    state_data["temp_uid"] = uid
    user_states[chat_id] = state_data
    
    return f"✅ UID: <code>{uid}</code>\n\n📦 Now send the payload (Cookie/2FA data):"

def handle_single_data(chat_id, payload):
    """State: AWAITING_SINGLE_DATA -> AWAITING_MANUAL_PASSWORD or AWAITING_UID"""
    state_data = user_states.get(chat_id, {})
    uid = state_data.get("temp_uid")
    category = state_data.get("category", "fb_cookie")
    
    if not uid:
        return "❌ Session expired. Please start over with /start."
        
    global_rule = get_setting("pass_rule", "")
    user_data = get_user_data(chat_id)
    custom_pass = user_data.get("custom_password", "")
    
    # Check if existing custom password satisfies the current global rule
    if validate_strict_password(custom_pass, global_rule):
        # Instant save
        track_id, rate = _save_submission(chat_id, uid, payload, category)
        if track_id:
            # Revert to AWAITING_UID for continuous fast submission
            state_data["state"] = "AWAITING_UID"
            # Clean up temp data to save RAM
            state_data.pop("temp_uid", None)
            user_states[chat_id] = state_data
            return f"✅ <b>Task Saved!</b>\n🎫 Track ID: <code>{track_id}</code>\n💰 Rate: {rate} BDT\n\n🔄 Send next UID:"
        return "❌ Duplicate submission detected."
    else:
        # Need manual password
        state_data["state"] = "AWAITING_MANUAL_PASSWORD"
        state_data["temp_payload"] = payload
        user_states[chat_id] = state_data
        return "🔐 Please enter your password to confirm this submission:"

def handle_manual_password(chat_id, password):
    """State: AWAITING_MANUAL_PASSWORD -> AWAITING_UID"""
    state_data = user_states.get(chat_id, {})
    uid = state_data.get("temp_uid")
    payload = state_data.get("temp_payload")
    category = state_data.get("category", "fb_cookie")
    
    if not uid or not payload:
        return "❌ Session expired. Please start over with /start."
        
    global_rule = get_setting("pass_rule", "")
    
    if validate_strict_password(password, global_rule):
        track_id, rate = _save_submission(chat_id, uid, payload, category)
        if track_id:
            # Save this password as the user's custom password for future instant saves
            update_user_field(chat_id, "custom_password", password)
            
            # Revert state
            state_data["state"] = "AWAITING_UID"
            # Clean up temp data to save RAM
            state_data.pop("temp_uid", None)
            state_data.pop("temp_payload", None)
            user_states[chat_id] = state_data
            
            return f"✅ <b>Task Saved!</b>\n🎫 Track ID: <code>{track_id}</code>\n💰 Rate: {rate} BDT\n🔓 Password saved for instant future submissions.\n\n🔄 Send next UID:"
        return "❌ Duplicate submission detected."
    else:
        return "❌ Incorrect password. Please try again or contact admin."

# ==========================================
# 3. Bulk Text Submission Flow
# ==========================================
def handle_bulk_text(chat_id, text):
    """State: AWAITING_BULK_TEXT -> Async Processing"""
    # 1. Strict Password Validation BEFORE processing
    global_rule = get_setting("pass_rule", "")
    user_data = get_user_data(chat_id)
    custom_pass = user_data.get("custom_password", "")
    
    if not validate_strict_password(custom_pass, global_rule):
        return "❌ Invalid Password. Bulk submission rejected. Set your password first."
        
    # 2. Send immediate feedback
    bot.send_message(chat_id, "⏳ <b>Processing bulk text...</b>\nThis may take a few moments.", parse_mode="HTML")
    
    # 3. Define the heavy background task
    def rbk():
        accepted_ids = []
        total_payout = 0.0
        cloud_blocked_count = 0
        
        try:
            lines = text.strip().split('\n')
            pending_list = []
            
            # Filter and prepare pending tasks
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                    
                uid = extract_numeric_uid(line)
                if not uid or is_duplicate_uid(uid):
                    continue
                    
                p_hash = generate_payload_hash(line)
                if is_payload_blacklisted(p_hash):
                    continue
                    
                # Determine category based on payload content
                category = "fb_cookie" if is_valid_cookies(line) else "fb_2fa"
                
                allowed, _ = is_submission_allowed(category, get_bd_time())
                if not allowed:
                    continue
                    
                pending_list.append({
                    "uid": uid,
                    "payload": line,
                    "category": category
                })
                
            if not pending_list:
                bot.send_message(chat_id, "❌ No valid, unique, and allowed tasks found in the text.")
                return

            # Concurrent Live Checking Wrapper
            def _chk(item):
                is_live, msg = check_live_account(item["uid"])
                item["is_live"] = is_live
                item["status_msg"] = msg
                return item

            # Map through the bounded live_check_executor
            checked_results = live_check_executor.map(_chk, pending_list)
            
            # Process results and save to DB
            for item in checked_results:
                if item["is_live"]:
                    track_id, rate = _save_submission(chat_id, item["uid"], item["payload"], item["category"])
                    if track_id:
                        accepted_ids.append(item["uid"])
                        total_payout += rate
                elif item["status_msg"] == "Cloud Blocked":
                    cloud_blocked_count += 1

            # 4. Send Final Summary
            summary = f"✅ <b>Bulk Processing Complete!</b>\n\n"
            summary += f"📥 <b>Accepted IDs:</b> {len(accepted_ids)}\n"
            summary += f"💰 <b>Total Payout:</b> {total_payout:.2f} BDT\n"
            
            if cloud_blocked_count > 0:
                summary += f"\n⚠️ <b>Warning:</b> {cloud_blocked_count} tasks were skipped due to Cloud IP Blocks. Please try again later."
                
            if accepted_ids:
                summary += f"\n📝 <b>First 5 IDs:</b>\n<code>{'\n'.join(accepted_ids[:5])}</code>"
                
            bot.send_message(chat_id, summary, parse_mode="HTML")
            
        except Exception as e:
            print(f"[BulkProcessor] Critical error: {e}")
            bot.send_message(chat_id, "❌ An internal error occurred during bulk processing.")
        finally:
            # CRITICAL: Force garbage collection to prevent RAM overflow on free-tier servers
            gc.collect()

    # Submit to the heavy task executor to keep the main thread unblocked
    heavy_task_executor.submit(rbk)
    
    return None  # Return None because the initial "Processing" message was already sent


# ==============================================================================
# OEB NEXUS - PART 5: Master File Processor (Excel/CSV Deep Extract)
# ==============================================================================
# DEPENDENCY: Assumes Parts 1-4 globals exist. Handles heavy file parsing, 
# concurrent live checking, atomic DB updates, and strict memory cleanup.

# ==========================================
# 1. Private Cloud Backup Function
# ==========================================
def send_private_backup_message(content, doc_buf=None, doc_name=None):
    """
    Sends a backup message and optional document to the private backup channel.
    Runs asynchronously to prevent blocking the main bot thread.
    """
    def task():
        try:
            # Truncate content to 3750 chars to strictly prevent Telegram API limits
            safe_content = str(content)[:3750]
            
            if doc_buf and doc_name:
                doc_buf.seek(0)
                bot.send_document(
                    BACKUP_CHANNEL_ID, 
                    doc_buf, 
                    caption=safe_content, 
                    parse_mode="HTML"
                )
            else:
                bot.send_message(
                    BACKUP_CHANNEL_ID, 
                    safe_content, 
                    parse_mode="HTML"
                )
        except Exception as e:
            print(f"[Backup] Failed to send backup: {e}")
            
    background_executor.submit(task)

# ==========================================
# 2. Document Handler & Router
# ==========================================
@bot.message_handler(content_types=['document'])
def handle_document(message):
    # Wrap execution in heavy_task_executor to protect main polling threads
    heavy_task_executor.submit(_process_document, message)

def _process_document(message):
    chat_id = message.chat.id
    
    # Basic security and state checks
    if get_setting("maintenance_mode", False):
        bot.send_message(chat_id, "🛠 Bot is under maintenance. Please try again later.")
        return
        
    if is_user_banned(chat_id):
        bot.send_message(chat_id, "🚫 You are banned from using this bot.")
        return
        
    state_data = user_states.get(chat_id, {})
    if state_data.get("state") != "AWAITING_EXCEL_FILE":
        bot.send_message(chat_id, "❌ Please use the menu to start an Excel/CSV submission first.")
        return

    # ==========================================
    # 3. Excel/CSV Deep Extraction
    # ==========================================
    # Strict password validation before downloading/processing
    global_rule = get_setting("pass_rule", "")
    user_data = get_user_data(chat_id)
    custom_pass = user_data.get("custom_password", "")
    
    if not validate_strict_password(custom_pass, global_rule):
        bot.send_message(chat_id, "❌ Invalid Password. Set your password first to process files.")
        return
        
    bot.send_message(chat_id, "⏳ <b>Processing file...</b>\nThis may take a few moments.", parse_mode="HTML")
    
    original_filename = message.document.file_name or "backup_file"
    file_ext = original_filename.lower().split('.')[-1]
    
    try:
        # Download file directly into RAM
        file_info = bot.get_file(message.document.file_id)
        dw = bot.download_file(file_info.file_path)
        
        cands = []
        
        if file_ext == "csv":
            # Decode and parse CSV
            text = dw.decode('utf-8', errors='ignore')
            reader = csv.reader(text.splitlines())
            for row in reader:
                uid = None
                payload = None
                for val in row:
                    val_str = str(val).strip()
                    if not val_str:
                        continue
                    if not uid:
                        u = extract_numeric_uid(val_str)
                        if u:
                            uid = u
                    if not payload:
                        # Ensure it's a valid cookie string OR a long non-numeric string
                        if is_valid_cookies(val_str) or (len(val_str) > 20 and not val_str.isdigit()):
                            payload = val_str
                if uid and payload and not is_duplicate_uid(uid):
                    cands.append({"uid": uid, "payload": payload})
                    
        elif file_ext in ["xlsx", "xls"]:
            # read_only=True and data_only=True are critical for strict memory efficiency
            wb = openpyxl.load_workbook(io.BytesIO(dw), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                uid = None
                payload = None
                for val in row:
                    if val is None:
                        continue
                    val_str = str(val).strip()
                    if not val_str:
                        continue
                    if not uid:
                        u = extract_numeric_uid(val_str)
                        if u:
                            uid = u
                    if not payload:
                        if is_valid_cookies(val_str) or (len(val_str) > 20 and not val_str.isdigit()):
                            payload = val_str
                if uid and payload and not is_duplicate_uid(uid):
                    cands.append({"uid": uid, "payload": payload})
            wb.close() # Explicitly close workbook to free memory
        else:
            bot.send_message(chat_id, "❌ Unsupported file format. Please send .csv or .xlsx files only.")
            return

        if not cands:
            bot.send_message(chat_id, "❌ No valid, unique tasks found in the file.")
            return

        # ==========================================
        # 4. Concurrent Validation & MongoDB Insertion
        # ==========================================
        valid_cands = []
        
        # Pre-filter candidates
        for c in cands:
            p_hash = generate_payload_hash(c["payload"])
            if is_payload_blacklisted(p_hash):
                continue
                
            category = "fb_cookie" if is_valid_cookies(c["payload"]) else "fb_2fa"
            allowed, _ = is_submission_allowed(category, get_bd_time())
            if not allowed:
                continue
                
            c["category"] = category
            valid_cands.append(c)
            
        if not valid_cands:
            bot.send_message(chat_id, "❌ All extracted tasks were blacklisted or past the daily deadline.")
            return

        # Concurrent Live Checking wrapper
        def _chk(item):
            is_live, msg = check_live_account(item["uid"])
            item["is_live"] = is_live
            item["status_msg"] = msg
            return item

        # Map through the bounded live_check_executor
        checked_results = live_check_executor.map(_chk, valid_cands)
        
        accepted_count = 0
        total_payout = 0.0
        
        # Process and save live accounts
        # Note: _save_submission (from Part 4) handles the atomic $inc for hold_balance
        for item in checked_results:
            if item["is_live"]:
                track_id, rate = _save_submission(chat_id, item["uid"], item["payload"], item["category"])
                if track_id:
                    accepted_count += 1
                    total_payout += rate

        # ==========================================
        # 5. Balances, Cloud Backup & Cleanup
        # ==========================================
        # Prepare backup file buffer (create a fresh BytesIO to avoid stream consumption issues)
        doc_buf = io.BytesIO(dw)
        backup_filename = f"Backup_{get_bd_time().strftime('%Y-%m-%d')}_{original_filename}"
        
        backup_caption = (
            f"📂 <b>FILE BACKUP</b>\n\n"
            f"👤 <b>Worker ID:</b> <code>{chat_id}</code>\n"
            f"🔑 <b>Password:</b> <code>{sanitize_html(custom_pass)}</code>\n"
            f"✅ <b>Valid Count:</b> {accepted_count}\n"
            f"💰 <b>Total Payout:</b> {total_payout:.2f} BDT\n"
            f"📄 <b>File:</b> {sanitize_html(original_filename)}"
        )
        
        send_private_backup_message(backup_caption, doc_buf=doc_buf, doc_name=backup_filename)
        
        # Send final summary to the worker
        summary = (
            f"✅ <b>File Processing Complete!</b>\n\n"
            f"📥 <b>Accepted Tasks:</b> {accepted_count}\n"
            f"💰 <b>Total Payout:</b> {total_payout:.2f} BDT\n\n"
            f"🔄 You can send another file or use /start to go to the main menu."
        )
        bot.send_message(chat_id, summary, parse_mode="HTML")
        
        # Revert state to allow continuous workflow
        state_data["state"] = "AWAITING_UID"
        user_states[chat_id] = state_data

    except Exception as e:
        print(f"[FileProcessor] Critical error: {e}")
        bot.send_message(chat_id, "❌ An internal error occurred while processing the file.")
    finally:
        # CRITICAL: Force garbage collection to prevent RAM overflow on 512MB servers
        gc.collect()


# ==============================================================================
# OEB NEXUS - PART 6: Auto-Matcher & Cloud Backup (Buyer Report Processing)
# ==============================================================================
# DEPENDENCY: Assumes Parts 1-5 globals exist. 
# NOTE: This part defines the processing function. The actual routing from the 
# document handler will be wired in Part 10 to avoid duplicate @bot.message_handler 
# decorators for 'document' content types.

# ==========================================
# 1. Auto-Matcher Processing Function
# ==========================================
def process_buyer_report(message):
    """
    Processes the buyer's report (Excel/CSV) uploaded by the Admin.
    Matches pending tasks, updates dual wallets atomically, and notifies workers.
    WARNING: This is a heavy task. MUST be executed via heavy_task_executor.
    """
    chat_id = message.chat.id
    
    # Strict Admin check
    if chat_id != ADMIN_ID:
        return
        
    state_data = user_states.get(chat_id, {})
    if state_data.get("state") != "AWAITING_BUYER_REPORT":
        return
        
    # Extract targets and clear state immediately to prevent re-triggering
    target_date = state_data.get("target_date", "ALL")
    target_cat = state_data.get("target_cat", "ALL")
    
    state_data["state"] = "AWAITING_UID"
    user_states[chat_id] = state_data
    
    bot.send_message(chat_id, "⏳ <b>Matching in progress...</b>\nPlease wait.", parse_mode="HTML")
    
    try:
        # ==========================================
        # 2. File Processing & UID Extraction
        # ==========================================
        file_info = bot.get_file(message.document.file_id)
        dw = bot.download_file(file_info.file_path)
        buf = io.BytesIO(dw)
        
        ex_uids = set()
        filename = message.document.file_name.lower() if message.document.file_name else ""
        
        if filename.endswith('.csv'):
            text = dw.decode('utf-8', errors='ignore')
            reader = csv.reader(text.splitlines())
            for row in reader:
                for val in row:
                    # Clean UIDs: remove decimals if parsed as floats by Excel
                    val_str = str(val).strip().replace('.0', '')
                    if val_str.isdigit() and 8 <= len(val_str) <= 20:
                        ex_uids.add(val_str)
                        
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            # read_only=True is critical for memory efficiency on large reports
            wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for val in row:
                    if val is None: 
                        continue
                    val_str = str(val).strip().replace('.0', '')
                    if val_str.isdigit() and 8 <= len(val_str) <= 20:
                        ex_uids.add(val_str)
            wb.close() # Explicitly close to free memory
        else:
            bot.send_message(chat_id, "❌ Unsupported file format. Please send .csv or .xlsx.")
            return

        if not ex_uids:
            bot.send_message(chat_id, "❌ No valid UIDs found in the buyer report.")
            return

        # ==========================================
        # 3. MongoDB Query & Matching
        # ==========================================
        q = {"status": "Hold"}
        if target_date != "ALL":
            q["date_key"] = target_date
        if target_cat != "ALL":
            q["category"] = target_cat
            
        subs = list(submissions_col.find(q))
        if not subs:
            bot.send_message(chat_id, "❌ No pending 'Hold' tasks found for the given criteria.")
            return
            
        appr = 0
        rej = 0
        payout = 0.0
        notifs = collections.defaultdict(list)
        
        # ==========================================
        # 4. Dual Wallet Payout Logic (Atomic Updates)
        # ==========================================
        for s in subs:
            uid = str(s.get("uid", ""))
            amt = float(s.get("rate", 0.0))
            wid = s.get("user_id")
            track_id = s.get("track_id")
            
            if not wid or not track_id:
                continue
                
            if uid in ex_uids:
                # --- APPROVED ---
                submissions_col.update_one({"_id": track_id}, {"$set": {"status": "Approved"}})
                
                # Fetch role to determine wallet destination
                user = users_col.find_one({"_id": wid}, {"role": 1})
                role = user.get("role", "member") if user else "member"
                
                # ATOMIC UPDATE: Use $gte to strictly prevent negative balances
                if role == "sub_admin":
                    users_col.update_one(
                        {"_id": wid, "hold_balance": {"$gte": amt}}, 
                        {"$inc": {"virtual_wallet": amt, "hold_balance": -amt}}
                    )
                else:
                    users_col.update_one(
                        {"_id": wid, "hold_balance": {"$gte": amt}}, 
                        {"$inc": {"balance": amt, "hold_balance": -amt}}
                    )
                    
                appr += 1
                payout += amt
                notifs[wid].append(f"✅ <b>Approved:</b> <code>{uid}</code> | 💰 +{amt} BDT")
                
            else:
                # --- REJECTED ---
                submissions_col.update_one({"_id": track_id}, {"$set": {"status": "Rejected"}})
                
                # ATOMIC UPDATE: Deduct hold balance without adding to main wallet
                users_col.update_one(
                    {"_id": wid, "hold_balance": {"$gte": amt}}, 
                    {"$inc": {"hold_balance": -amt}}
                )
                
                rej += 1
                notifs[wid].append(f"❌ <b>Rejected:</b> <code>{uid}</code>")

        # ==========================================
        # 5. Anti-Spam Batch Notifications & Cleanup
        # ==========================================
        for wid, msgs in notifs.items():
            # Limit to 15 lines to prevent oversized messages and API errors
            if len(msgs) > 15:
                text = "\n".join(msgs[:15]) + f"\n<i>...and {len(msgs) - 15} more.</i>"
            else:
                text = "\n".join(msgs)
                
            try:
                bot.send_message(wid, text, parse_mode="HTML")
                # STRICT ANTI-SPAM DELAY to prevent 429 Too Many Requests
                time.sleep(0.04) 
            except Exception as e:
                print(f"[Notifier] Failed to send to {wid}: {e}")
                
        # Final summary to Admin
        summary = (
            f"✅ <b>Auto-Matcher Complete!</b>\n\n"
            f"📊 <b>Total Approved:</b> {appr}\n"
            f"💰 <b>Total Payout:</b> {payout:.2f} BDT\n"
            f"❌ <b>Total Rejected:</b> {rej}\n"
            f"👥 <b>Workers Notified:</b> {len(notifs)}"
        )
        bot.send_message(ADMIN_ID, summary, parse_mode="HTML")
        
    except Exception as e:
        print(f"[AutoMatcher] Critical error: {e}")
        bot.send_message(chat_id, "❌ An internal error occurred during matching.")
    finally:
        # CRITICAL: Force garbage collection to release heavy Excel/CSV processing memory
        gc.collect()


# ==============================================================================
# OEB NEXUS - PART 7: Dynamic UI, Main Router, Callbacks & Final Execution
# ==============================================================================
# DEPENDENCY: Assumes Parts 1-6 globals exist. 
# This part wires the entire user interface, routes text/callbacks, and starts the bot.

# ==========================================
# 1. Dynamic Keyboard Generators
# ==========================================
def main_bottom_keyboard(chat_id):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("Submit Tasks", "Helper Tools", "Profile & Wallet", "Reward & Support")
    user = get_user_data(chat_id)
    if chat_id == ADMIN_ID or user.get("role") in ["admin", "sub_admin"]:
        kb.add("Admin Control Center")
    return kb

def submit_tasks_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("Single Submission", "Bulk Text", "Excel/CSV File", "Back")
    return kb

def helper_tools_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("Temp Email", "2FA Generator", "Leaderboard", "Back")
    return kb

def account_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("My Balance", "Withdraw", "Transaction History", "Back")
    return kb

def bonus_support_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("Daily Bonus", "Contact Support", "Back")
    return kb

def cancel_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("❌ Cancel")
    return kb

def category_bottom_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    rates = get_setting("rates", {"fb_cookie": 5.0, "fb_2fa": 6.0, "ig_cookie": 8.0, "ig_2fa": 10.0})
    surge = get_active_surge_bonus()
    for cat, base in rates.items():
        total = base + surge
        kb.add(f"{cat} ({total} BDT)")
    kb.add("Back")
    return kb

def admin_bottom_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("Task Management", "Finance & Wallets", "Settings", "System Tools", "Main Menu")
    return kb

def admin_sub_task_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("Buyer Report", "Force Close Shift", "Export Data", "Back")
    return kb

def admin_sub_finance_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("Pending Withdrawals", "Adjust Balance", "Back")
    return kb

def admin_sub_settings_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("Password Rule", "Base Rates", "Surge Pricing", "Back")
    return kb

def admin_sub_system_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add("Broadcast Message", "Maintenance Mode", "Back")
    return kb

# ==========================================
# 2. Start Command & Main Menu
# ==========================================
@bot.message_handler(commands=['start', 'menu', 'help'])
def cmd_start(message):
    chat_id = message.chat.id
    user_states.pop(chat_id, None) # Clear any orphaned states
    
    if not check_force_join(chat_id):
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton("✅ Verify Join", callback_data="verify_join"))
        msg = "⚠️ <b>Force Join Required!</b>\n\nPlease join the following channels:\n"
        for ch in REQUIRED_CHANNELS:
            msg += f"➡️ {ch['name']}: {ch['url']}\n"
        bot.send_message(chat_id, msg, parse_mode="HTML", reply_markup=kb)
        return

    bot.send_message(
        chat_id, 
        f"🏠 <b>Welcome to OEB NEXUS!</b>\n\nSelect an option below:", 
        parse_mode="HTML", 
        reply_markup=main_bottom_keyboard(chat_id)
    )

# ==========================================
# 3. Main Text Router
# ==========================================
@bot.message_handler(content_types=['text'])
def handle_text(message):
    # Route heavy processing to background executor to protect polling thread
    background_executor.submit(_process_main_router, message)

def _process_main_router(message):
    chat_id = message.chat.id
    text = message.text.strip()
    
    # Security & Maintenance Checks
    if get_setting("maintenance_mode", False) and chat_id != ADMIN_ID:
        bot.send_message(chat_id, "🛠 Bot is under maintenance. Please try again later.")
        return
        
    if is_user_banned(chat_id):
        bot.send_message(chat_id, "🚫 You are banned from using this bot.")
        return

    # State Reset for Navigation
    if text in ["Main Menu", "Back", "Cancel", "🏠 Main Menu", "❌ Cancel"]:
        user_states.pop(chat_id, None)
        bot.send_message(chat_id, "🏠 <b>Main Menu</b>", parse_mode="HTML", reply_markup=main_bottom_keyboard(chat_id))
        return

    # State Handling (Routing to specific flows)
    state_data = user_states.get(chat_id)
    if state_data:
        state = state_data.get("state")
        
        if state == "AWAITING_UID":
            bot.send_message(chat_id, handle_single_uid(chat_id, text), parse_mode="HTML", reply_markup=cancel_keyboard())
            return
        elif state == "AWAITING_SINGLE_DATA":
            bot.send_message(chat_id, handle_single_data(chat_id, text), parse_mode="HTML", reply_markup=cancel_keyboard())
            return
        elif state == "AWAITING_MANUAL_PASSWORD":
            bot.send_message(chat_id, handle_manual_password(chat_id, text), parse_mode="HTML", reply_markup=cancel_keyboard())
            return
        elif state == "AWAITING_BULK_TEXT":
            res = handle_bulk_text(chat_id, text)
            if res: bot.send_message(chat_id, res, parse_mode="HTML")
            return
        elif state == "AWAITING_2FA_GEN":
            secret = text.replace(" ", "").upper()
            try:
                totp = pyotp.TOTP(secret)
                code = totp.now()
                bot.send_message(chat_id, f"🔐 <b>2FA Code:</b> <code>{code}</code>\n\n<i>Valid for 30 seconds.</i>", parse_mode="HTML", reply_markup=helper_tools_keyboard())
            except Exception:
                bot.send_message(chat_id, "❌ Invalid Secret Key.", reply_markup=helper_tools_keyboard())
            user_states.pop(chat_id, None)
            return
        elif state == "AWAITING_ADMIN_PASS_RULE" and chat_id == ADMIN_ID:
            update_setting("pass_rule", text)
            bot.send_message(chat_id, f"✅ Password rule updated to: <code>{text}</code>", parse_mode="HTML", reply_markup=admin_sub_settings_keyboard())
            user_states.pop(chat_id, None)
            return
        elif state == "AWAITING_BROADCAST_MSG" and chat_id == ADMIN_ID:
            bot.send_message(chat_id, "🔄 Broadcasting...")
            users = list(users_col.find({"banned": False}))
            count = 0
            for u in users:
                try:
                    bot.send_message(u["_id"], text, parse_mode="HTML")
                    count += 1
                    time.sleep(0.04)
                except Exception: pass
            bot.send_message(chat_id, f"✅ Broadcasted to {count} users.", reply_markup=admin_sub_system_keyboard())
            user_states.pop(chat_id, None)
            return

    # Menu Routing (Top Level)
    if text == "Submit Tasks":
        bot.send_message(chat_id, "📦 <b>Select Submission Type:</b>", parse_mode="HTML", reply_markup=submit_tasks_keyboard())
    elif text == "Helper Tools":
        bot.send_message(chat_id, "🛠 <b>Helper Tools:</b>", parse_mode="HTML", reply_markup=helper_tools_keyboard())
    elif text == "Profile & Wallet":
        bot.send_message(chat_id, "💼 <b>Profile & Wallet:</b>", parse_mode="HTML", reply_markup=account_keyboard())
    elif text == "Reward & Support":
        bot.send_message(chat_id, "🎁 <b>Rewards & Support:</b>", parse_mode="HTML", reply_markup=bonus_support_keyboard())
    elif text == "Admin Control Center":
        if chat_id == ADMIN_ID or get_user_data(chat_id).get("role") in ["admin", "sub_admin"]:
            bot.send_message(chat_id, "⚙️ <b>Admin Control Center:</b>", parse_mode="HTML", reply_markup=admin_bottom_keyboard())
            
    # Submission Sub-menus
    elif text == "Single Submission":
        bot.send_message(chat_id, "📦 <b>Select Category:</b>", parse_mode="HTML", reply_markup=category_bottom_keyboard())
        # Note: In a full implementation, clicking a category button would set the state to AWAITING_UID and store the category.
    elif text == "Bulk Text":
        state_data = user_states.get(chat_id, {})
        state_data["state"] = "AWAITING_BULK_TEXT"
        user_states[chat_id] = state_data
        bot.send_message(chat_id, "📝 Send your bulk text (one task per line):", reply_markup=cancel_keyboard())
    elif text == "Excel/CSV File":
        state_data = user_states.get(chat_id, {})
        state_data["state"] = "AWAITING_EXCEL_FILE"
        user_states[chat_id] = state_data
        bot.send_message(chat_id, "📂 Send your .xlsx or .csv file:", reply_markup=cancel_keyboard())

    # Helper Tools Actions
    elif text == "Temp Email":
        rand_str = ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
        email = f"{rand_str}@1secmail.com"
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton("📥 Check Inbox", callback_data=f"check_otp_{email}"))
        bot.send_message(chat_id, f"📧 <b>Your Temp Email:</b>\n<code>{email}</code>", parse_mode="HTML", reply_markup=kb)
    elif text == "2FA Generator":
        bot.send_message(chat_id, "🔐 Send the <b>2FA Secret Key</b>:", parse_mode="HTML", reply_markup=cancel_keyboard())
        state_data = user_states.get(chat_id, {})
        state_data["state"] = "AWAITING_2FA_GEN"
        user_states[chat_id] = state_data
    elif text == "Leaderboard":
        pipeline = [
            {"$group": {"_id": "$user_id", "total_earnings": {"$sum": "$rate"}, "total_tasks": {"$sum": 1}}},
            {"$sort": {"total_earnings": -1}}, {"$limit": 10}
        ]
        results = list(submissions_col.aggregate(pipeline))
        msg = "🏆 <b>Top 10 Workers</b>\n\n"
        for i, r in enumerate(results, 1):
            msg += f"{i}. <code>{r['_id']}</code> - 💰 {r['total_earnings']:.2f} BDT ({r['total_tasks']} tasks)\n"
        bot.send_message(chat_id, msg, parse_mode="HTML", reply_markup=helper_tools_keyboard())

    # Admin Sub-menus & Actions
    elif text == "Task Management":
        bot.send_message(chat_id, "📋 <b>Task Management:</b>", parse_mode="HTML", reply_markup=admin_sub_task_keyboard())
    elif text == "Finance & Wallets":
        bot.send_message(chat_id, "💰 <b>Finance & Wallets:</b>", parse_mode="HTML", reply_markup=admin_sub_finance_keyboard())
    elif text == "Settings":
        bot.send_message(chat_id, "⚙️ <b>Settings:</b>", parse_mode="HTML", reply_markup=admin_sub_settings_keyboard())
    elif text == "System Tools":
        bot.send_message(chat_id, "🖥 <b>System Tools:</b>", parse_mode="HTML", reply_markup=admin_sub_system_keyboard())
    elif text == "Password Rule":
        bot.send_message(chat_id, "🔑 Send the new password rule (or 'none' to disable):", reply_markup=cancel_keyboard())
        state_data = user_states.get(chat_id, {})
        state_data["state"] = "AWAITING_ADMIN_PASS_RULE"
        user_states[chat_id] = state_data
    elif text == "Broadcast Message":
        bot.send_message(chat_id, "📢 Send the message to broadcast:", reply_markup=cancel_keyboard())
        state_data = user_states.get(chat_id, {})
        state_data["state"] = "AWAITING_BROADCAST_MSG"
        user_states[chat_id] = state_data
    elif text == "Buyer Report":
        if chat_id == ADMIN_ID:
            bot.send_message(chat_id, "📊 Send the buyer report (.csv/.xlsx) to match 'Hold' tasks.", reply_markup=cancel_keyboard())
            state_data = user_states.get(chat_id, {})
            state_data["state"] = "AWAITING_BUYER_REPORT"
            state_data["target_date"] = "ALL"
            state_data["target_cat"] = "ALL"
            user_states[chat_id] = state_data

# ==========================================
# 4. Callback Query Handler
# ==========================================
@bot.callback_query_handler(func=lambda call: True)
def handle_callback(call):
    background_executor.submit(_process_callbacks, call)

def _process_callbacks(call):
    chat_id = call.message.chat.id
    data = call.data
    
    try:
        # 1. Force Join Verification
        if data == "verify_join":
            if check_force_join(chat_id):
                bot.answer_callback_query(call.id, "✅ Verified Successfully!")
                bot.send_message(chat_id, "🏠 <b>Main Menu</b>", parse_mode="HTML", reply_markup=main_bottom_keyboard(chat_id))
            else:
                bot.answer_callback_query(call.id, "❌ You haven't joined the channels yet!", show_alert=True)

        # 2. Temp Mail OTP Check
        elif data.startswith("check_otp_"):
            email = data.replace("check_otp_", "")
            bot.answer_callback_query(call.id, "🔄 Fetching inbox...")
            
            username, domain = email.split('@')
            url = f"https://www.1secmail.com/api/v1/?action=getMessages&login={username}&domain={domain}"
            
            res = requests.get(url, timeout=5.0).json()
            if not res:
                bot.send_message(chat_id, "📭 Inbox is empty. Try again in a few seconds.", reply_markup=helper_tools_keyboard())
                return
                
            msg_id = res[-1]['id']
            url2 = f"https://www.1secmail.com/api/v1/?action=readMessage&login={username}&domain={domain}&id={msg_id}"
            msg_data = requests.get(url2, timeout=5.0).json()
            
            body = msg_data.get('textBody', '') or msg_data.get('body', '')
            safe_body = sanitize_html(body)
            
            bot.send_message(chat_id, f"📩 <b>Latest Email Body:</b>\n\n<pre>{safe_body}</pre>", parse_mode="HTML", reply_markup=helper_tools_keyboard())

        # 3. Admin Dashboard Tools
        elif data == "shift_next_day":
            if chat_id != ADMIN_ID: return
            tomorrow = (get_bd_time() + timedelta(days=1)).strftime("%Y-%m-%d")
            config = get_shift_config()
            config["current_date"] = tomorrow
            update_setting("shift_config", config)
            bot.answer_callback_query(call.id, f"✅ Shift advanced to {tomorrow}")
            
        elif data.startswith("force_close_"):
            if chat_id != ADMIN_ID: return
            date_str = data.replace("force_close_", "")
            bot.answer_callback_query(call.id, "🔄 Force closing shift...")
            
            subs = submissions_col.find({"date_key": date_str, "status": "Hold"})
            count = 0
            for s in subs:
                amt = float(s.get("rate", 0.0))
                wid = s.get("user_id")
                submissions_col.update_one({"_id": s["_id"]}, {"$set": {"status": "Rejected"}})
                users_col.update_one({"_id": wid, "hold_balance": {"$gte": amt}}, {"$inc": {"hold_balance": -amt}})
                count += 1
                
            bot.send_message(chat_id, f"✅ Force closed {count} tasks for {date_str}. Escrow refunded.", reply_markup=admin_sub_task_keyboard())
            gc.collect()
            
        elif data.startswith("exp_select_date_"):
            if chat_id != ADMIN_ID: return
            date_str = data.replace("exp_select_date_", "")
            bot.answer_callback_query(call.id, "🔄 Generating CSV...")
            
            subs = submissions_col.find({"date_key": date_str})
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(["Track ID", "UID", "Category", "Rate", "Status", "User ID"])
            for s in subs:
                writer.writerow([s.get('track_id'), s.get('uid'), s.get('category'), s.get('rate'), s.get('status'), s.get('user_id')])
                
            buf.seek(0)
            doc_buf = io.BytesIO(buf.getvalue().encode('utf-8'))
            bot.send_document(chat_id, doc_buf, caption=f"📊 Report for {date_str}", visible_file_name=f"Report_{date_str}.csv", reply_markup=admin_sub_task_keyboard())
            gc.collect()

    except Exception as e:
        print(f"[Callback] Critical error: {e}")
        bot.answer_callback_query(call.id, "❌ An error occurred.", show_alert=True)

# ==========================================
# 5. Final Execution Block (Start Server & Bot)
# ==========================================
if __name__ == "__main__":
    print("="*50)
    print("🚀 OEB NEXUS ENGINE STARTING...")
    print("="*50)
    
    # Start the 24/7 Web Server in a daemon thread (keeps Render/Heroku alive)
    web_thread = threading.Thread(target=run_web, daemon=True)
    web_thread.start()
    print("[Server] Flask + Waitress running on port 10000.")
    
    # Start the Telegram Bot Polling
    print("[Bot] Starting infinity polling...")
    # skip_pending=True ensures we don't process old messages and crash RAM on startup
    bot.infinity_polling(skip_pending=True, timeout=60, long_polling_timeout=60)


